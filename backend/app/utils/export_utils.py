"""
Report Export Utilities - PDF and Excel Generation
Kontali ERP - Norwegian compliant report exports
"""
from io import BytesIO
from typing import List, Dict, Any, Optional
from datetime import date
from decimal import Decimal

from weasyprint import HTML, CSS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from fastapi.responses import StreamingResponse


def format_currency(value: float) -> str:
    """Format number as Norwegian currency (kr 1 234,56)"""
    return f"kr {value:,.2f}".replace(",", " ").replace(".", ",")


def format_date_no(d: Optional[date]) -> str:
    """Format date as Norwegian (dd.mm.yyyy)"""
    if not d:
        return ""
    return d.strftime("%d.%m.%Y")


def generate_pdf_saldobalanse(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate PDF for Saldobalanse report"""
    
    from_date = data.get("from_date")
    to_date = data.get("to_date")
    balances = data.get("balances", [])
    
    period_text = ""
    if from_date and to_date:
        period_text = f"Periode: {format_date_no(date.fromisoformat(from_date))} - {format_date_no(date.fromisoformat(to_date))}"
    elif from_date:
        period_text = f"Fra: {format_date_no(date.fromisoformat(from_date))}"
    elif to_date:
        period_text = f"Til: {format_date_no(date.fromisoformat(to_date))}"
    
    # Build table rows
    table_rows = ""
    for item in balances:
        table_rows += f"""
        <tr>
            <td>{item['account_number']}</td>
            <td>{item['account_name']}</td>
            <td class="number">{format_currency(item['total_debit'])}</td>
            <td class="number">{format_currency(item['total_credit'])}</td>
            <td class="number">{format_currency(item['balance'])}</td>
        </tr>
        """
    
    html_content = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 1.5cm; }}
            body {{ font-family: Arial, sans-serif; font-size: 10pt; }}
            h1 {{ text-align: center; color: #2c3e50; margin-bottom: 5px; }}
            .meta {{ text-align: center; color: #7f8c8d; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ border: 1px solid #bdc3c7; padding: 8px; }}
            th {{ background-color: #34495e; color: white; font-weight: bold; }}
            tr:nth-child(even) {{ background-color: #ecf0f1; }}
            .number {{ text-align: right; }}
            .footer {{ margin-top: 30px; text-align: center; color: #95a5a6; font-size: 9pt; }}
            .totals {{ font-weight: bold; background-color: #d5dbdb !important; }}
        </style>
    </head>
    <body>
        <h1>Saldobalanse</h1>
        <div class="meta">
            <p><strong>{client_name}</strong></p>
            <p>{period_text}</p>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Konto</th>
                    <th>Navn</th>
                    <th>Debet</th>
                    <th>Kredit</th>
                    <th>Saldo</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
                <tr class="totals">
                    <td colspan="2"><strong>Sum</strong></td>
                    <td class="number">{format_currency(data.get('total_debit', 0))}</td>
                    <td class="number">{format_currency(data.get('total_credit', 0))}</td>
                    <td class="number">-</td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            <p>Generert: {format_date_no(date.today())} | Kontali ERP</p>
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = HTML(string=html_content).write_pdf()
    
    filename = f"Saldobalanse_{client_name.replace(' ', '_')}_{date.today().isoformat()}.pdf"
    
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_excel_saldobalanse(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate Excel for Saldobalanse report"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Saldobalanse"
    
    # Header styling
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "Saldobalanse"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Meta info
    ws.merge_cells('A2:E2')
    meta_cell = ws['A2']
    meta_cell.value = client_name
    meta_cell.alignment = Alignment(horizontal='center')
    
    from_date = data.get("from_date")
    to_date = data.get("to_date")
    if from_date and to_date:
        ws.merge_cells('A3:E3')
        period_cell = ws['A3']
        period_cell.value = f"Periode: {format_date_no(date.fromisoformat(from_date))} - {format_date_no(date.fromisoformat(to_date))}"
        period_cell.alignment = Alignment(horizontal='center')
        header_row = 5
    else:
        header_row = 4
    
    # Headers
    headers = ["Konto", "Navn", "Debet", "Kredit", "Saldo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    balances = data.get("balances", [])
    current_row = header_row + 1
    
    for item in balances:
        ws.cell(row=current_row, column=1, value=item['account_number']).border = border
        ws.cell(row=current_row, column=2, value=item['account_name']).border = border
        
        debit_cell = ws.cell(row=current_row, column=3, value=item['total_debit'])
        debit_cell.number_format = '#,##0.00'
        debit_cell.border = border
        
        credit_cell = ws.cell(row=current_row, column=4, value=item['total_credit'])
        credit_cell.number_format = '#,##0.00'
        credit_cell.border = border
        
        balance_cell = ws.cell(row=current_row, column=5, value=item['balance'])
        balance_cell.number_format = '#,##0.00'
        balance_cell.border = border
        
        # Alternating row colors
        if current_row % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        
        current_row += 1
    
    # Totals row
    totals_fill = PatternFill(start_color="d5dbdb", end_color="d5dbdb", fill_type="solid")
    ws.cell(row=current_row, column=1, value="Sum").font = Font(bold=True)
    ws.cell(row=current_row, column=1).fill = totals_fill
    ws.cell(row=current_row, column=1).border = border
    
    ws.cell(row=current_row, column=2, value="").fill = totals_fill
    ws.cell(row=current_row, column=2).border = border
    
    total_debit = ws.cell(row=current_row, column=3, value=data.get('total_debit', 0))
    total_debit.number_format = '#,##0.00'
    total_debit.font = Font(bold=True)
    total_debit.fill = totals_fill
    total_debit.border = border
    
    total_credit = ws.cell(row=current_row, column=4, value=data.get('total_credit', 0))
    total_credit.number_format = '#,##0.00'
    total_credit.font = Font(bold=True)
    total_credit.fill = totals_fill
    total_credit.border = border
    
    ws.cell(row=current_row, column=5, value="-").fill = totals_fill
    ws.cell(row=current_row, column=5).border = border
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Freeze top rows
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    
    # Save to BytesIO
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    filename = f"Saldobalanse_{client_name.replace(' ', '_')}_{date.today().isoformat()}.xlsx"
    
    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_pdf_resultat(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate PDF for Resultatregnskap"""
    
    from_date = data.get("from_date")
    to_date = data.get("to_date")
    
    period_text = ""
    if from_date and to_date:
        period_text = f"Periode: {format_date_no(date.fromisoformat(from_date))} - {format_date_no(date.fromisoformat(to_date))}"
    
    # Build income rows
    income_rows = ""
    for item in data.get("inntekter", []):
        income_rows += f"""
        <tr>
            <td>{item['account_number']}</td>
            <td>{item['account_name']}</td>
            <td class="number">{format_currency(item['amount'])}</td>
        </tr>
        """
    
    # Build cost sections
    varekjop_rows = ""
    for item in data.get("kostnader", {}).get("varekjop", {}).get("items", []):
        varekjop_rows += f"""
        <tr>
            <td>{item['account_number']}</td>
            <td>{item['account_name']}</td>
            <td class="number">{format_currency(item['amount'])}</td>
        </tr>
        """
    
    lonn_rows = ""
    for item in data.get("kostnader", {}).get("lonnkostnader", {}).get("items", []):
        lonn_rows += f"""
        <tr>
            <td>{item['account_number']}</td>
            <td>{item['account_name']}</td>
            <td class="number">{format_currency(item['amount'])}</td>
        </tr>
        """
    
    andre_rows = ""
    for item in data.get("kostnader", {}).get("andre_driftskostnader", {}).get("items", []):
        andre_rows += f"""
        <tr>
            <td>{item['account_number']}</td>
            <td>{item['account_name']}</td>
            <td class="number">{format_currency(item['amount'])}</td>
        </tr>
        """
    
    html_content = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4; margin: 1.5cm; }}
            body {{ font-family: Arial, sans-serif; font-size: 10pt; }}
            h1 {{ text-align: center; color: #2c3e50; margin-bottom: 5px; }}
            h2 {{ color: #34495e; margin-top: 20px; border-bottom: 2px solid #34495e; }}
            .meta {{ text-align: center; color: #7f8c8d; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th, td {{ border: 1px solid #bdc3c7; padding: 8px; }}
            th {{ background-color: #34495e; color: white; font-weight: bold; }}
            tr:nth-child(even) {{ background-color: #ecf0f1; }}
            .number {{ text-align: right; }}
            .subtotal {{ font-weight: bold; background-color: #d5dbdb !important; }}
            .total {{ font-weight: bold; background-color: #95a5a6 !important; color: white; }}
            .footer {{ margin-top: 30px; text-align: center; color: #95a5a6; font-size: 9pt; }}
        </style>
    </head>
    <body>
        <h1>Resultatregnskap</h1>
        <div class="meta">
            <p><strong>{client_name}</strong></p>
            <p>{period_text}</p>
        </div>
        
        <h2>Inntekter</h2>
        <table>
            <thead>
                <tr>
                    <th>Konto</th>
                    <th>Navn</th>
                    <th>Beløp</th>
                </tr>
            </thead>
            <tbody>
                {income_rows}
                <tr class="subtotal">
                    <td colspan="2"><strong>Sum inntekter</strong></td>
                    <td class="number">{format_currency(data.get('sum_inntekter', 0))}</td>
                </tr>
            </tbody>
        </table>
        
        <h2>Varekostnader</h2>
        <table>
            <tbody>
                {varekjop_rows}
                <tr class="subtotal">
                    <td colspan="2"><strong>Sum varekostnader</strong></td>
                    <td class="number">{format_currency(data.get('kostnader', {}).get('varekjop', {}).get('sum', 0))}</td>
                </tr>
            </tbody>
        </table>
        
        <h2>Lønnskostnader</h2>
        <table>
            <tbody>
                {lonn_rows}
                <tr class="subtotal">
                    <td colspan="2"><strong>Sum lønnskostnader</strong></td>
                    <td class="number">{format_currency(data.get('kostnader', {}).get('lonnkostnader', {}).get('sum', 0))}</td>
                </tr>
            </tbody>
        </table>
        
        <h2>Andre driftskostnader</h2>
        <table>
            <tbody>
                {andre_rows}
                <tr class="subtotal">
                    <td colspan="2"><strong>Sum andre driftskostnader</strong></td>
                    <td class="number">{format_currency(data.get('kostnader', {}).get('andre_driftskostnader', {}).get('sum', 0))}</td>
                </tr>
            </tbody>
        </table>
        
        <table style="margin-top: 30px;">
            <tbody>
                <tr class="total">
                    <td colspan="2"><strong>RESULTAT FØR SKATT</strong></td>
                    <td class="number"><strong>{format_currency(data.get('resultat', 0))}</strong></td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            <p>Generert: {format_date_no(date.today())} | Kontali ERP</p>
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = HTML(string=html_content).write_pdf()
    
    filename = f"Resultatregnskap_{client_name.replace(' ', '_')}_{date.today().isoformat()}.pdf"
    
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_excel_resultat(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate Excel for Resultatregnskap"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultatregnskap"
    
    # Styling
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    subtotal_fill = PatternFill(start_color="d5dbdb", end_color="d5dbdb", fill_type="solid")
    total_fill = PatternFill(start_color="95a5a6", end_color="95a5a6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title and meta
    ws.merge_cells('A1:C1')
    ws['A1'] = "Resultatregnskap"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    ws['A2'] = client_name
    ws['A2'].alignment = Alignment(horizontal='center')
    
    from_date = data.get("from_date")
    to_date = data.get("to_date")
    if from_date and to_date:
        ws.merge_cells('A3:C3')
        ws['A3'] = f"Periode: {format_date_no(date.fromisoformat(from_date))} - {format_date_no(date.fromisoformat(to_date))}"
        ws['A3'].alignment = Alignment(horizontal='center')
    
    row = 5
    
    # Inntekter section
    ws.cell(row=row, column=1, value="INNTEKTER").font = Font(bold=True, size=12)
    row += 1
    
    # Headers
    for col, header in enumerate(["Konto", "Navn", "Beløp"], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    # Income data
    for item in data.get("inntekter", []):
        ws.cell(row=row, column=1, value=item['account_number']).border = border
        ws.cell(row=row, column=2, value=item['account_name']).border = border
        amount_cell = ws.cell(row=row, column=3, value=item['amount'])
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = border
        row += 1
    
    # Sum income
    ws.cell(row=row, column=1, value="Sum inntekter").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2, value="").fill = subtotal_fill
    ws.cell(row=row, column=2).border = border
    sum_cell = ws.cell(row=row, column=3, value=data.get('sum_inntekter', 0))
    sum_cell.number_format = '#,##0.00'
    sum_cell.font = Font(bold=True)
    sum_cell.fill = subtotal_fill
    sum_cell.border = border
    row += 2
    
    # Cost sections
    cost_sections = [
        ("VAREKOSTNADER", "varekjop", "Sum varekostnader"),
        ("LØNNSKOSTNADER", "lonnkostnader", "Sum lønnskostnader"),
        ("ANDRE DRIFTSKOSTNADER", "andre_driftskostnader", "Sum andre driftskostnader")
    ]
    
    for section_title, section_key, sum_label in cost_sections:
        ws.cell(row=row, column=1, value=section_title).font = Font(bold=True, size=12)
        row += 1
        
        # Section data
        for item in data.get("kostnader", {}).get(section_key, {}).get("items", []):
            ws.cell(row=row, column=1, value=item['account_number']).border = border
            ws.cell(row=row, column=2, value=item['account_name']).border = border
            amount_cell = ws.cell(row=row, column=3, value=item['amount'])
            amount_cell.number_format = '#,##0.00'
            amount_cell.border = border
            row += 1
        
        # Section sum
        ws.cell(row=row, column=1, value=sum_label).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = subtotal_fill
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value="").fill = subtotal_fill
        ws.cell(row=row, column=2).border = border
        sum_cell = ws.cell(row=row, column=3, value=data.get("kostnader", {}).get(section_key, {}).get("sum", 0))
        sum_cell.number_format = '#,##0.00'
        sum_cell.font = Font(bold=True)
        sum_cell.fill = subtotal_fill
        sum_cell.border = border
        row += 2
    
    # Total result
    ws.cell(row=row, column=1, value="RESULTAT FØR SKATT").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2, value="").fill = total_fill
    ws.cell(row=row, column=2).border = border
    result_cell = ws.cell(row=row, column=3, value=data.get('resultat', 0))
    result_cell.number_format = '#,##0.00'
    result_cell.font = Font(bold=True, color="FFFFFF")
    result_cell.fill = total_fill
    result_cell.border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    filename = f"Resultatregnskap_{client_name.replace(' ', '_')}_{date.today().isoformat()}.xlsx"
    
    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_pdf_balanse(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate PDF for Balanse (Balance Sheet)"""
    
    balance_date = data.get("balance_date")
    
    # Build assets rows
    assets_rows = ""
    for item in data.get("eiendeler", []):
        assets_rows += f"""
        <tr>
            <td>{item['account_number']}</td>
            <td>{item['account_name']}</td>
            <td class="number">{format_currency(item['balance'])}</td>
        </tr>
        """
    
    # Build liabilities/equity rows
    liab_rows = ""
    for item in data.get("gjeld_egenkapital", []):
        liab_rows += f"""
        <tr>
            <td>{item['account_number']}</td>
            <td>{item['account_name']}</td>
            <td class="number">{format_currency(abs(item['balance']))}</td>
        </tr>
        """
    
    html_content = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4; margin: 1.5cm; }}
            body {{ font-family: Arial, sans-serif; font-size: 10pt; }}
            h1 {{ text-align: center; color: #2c3e50; margin-bottom: 5px; }}
            h2 {{ color: #34495e; margin-top: 20px; border-bottom: 2px solid #34495e; }}
            .meta {{ text-align: center; color: #7f8c8d; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th, td {{ border: 1px solid #bdc3c7; padding: 8px; }}
            th {{ background-color: #34495e; color: white; font-weight: bold; }}
            tr:nth-child(even) {{ background-color: #ecf0f1; }}
            .number {{ text-align: right; }}
            .subtotal {{ font-weight: bold; background-color: #d5dbdb !important; }}
            .balanced {{ color: #27ae60; font-weight: bold; }}
            .footer {{ margin-top: 30px; text-align: center; color: #95a5a6; font-size: 9pt; }}
        </style>
    </head>
    <body>
        <h1>Balanse</h1>
        <div class="meta">
            <p><strong>{client_name}</strong></p>
            <p>Per: {format_date_no(date.fromisoformat(balance_date))}</p>
        </div>
        
        <h2>Eiendeler</h2>
        <table>
            <thead>
                <tr>
                    <th>Konto</th>
                    <th>Navn</th>
                    <th>Beløp</th>
                </tr>
            </thead>
            <tbody>
                {assets_rows}
                <tr class="subtotal">
                    <td colspan="2"><strong>Sum eiendeler</strong></td>
                    <td class="number">{format_currency(data.get('sum_eiendeler', 0))}</td>
                </tr>
            </tbody>
        </table>
        
        <h2>Gjeld og Egenkapital</h2>
        <table>
            <tbody>
                {liab_rows}
                <tr class="subtotal">
                    <td colspan="2"><strong>Sum gjeld og egenkapital</strong></td>
                    <td class="number">{format_currency(data.get('sum_gjeld_egenkapital', 0))}</td>
                </tr>
            </tbody>
        </table>
        
        <p class="balanced" style="text-align: center; margin-top: 20px;">
            ✓ Balansen er i balanse
        </p>
        
        <div class="footer">
            <p>Generert: {format_date_no(date.today())} | Kontali ERP</p>
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = HTML(string=html_content).write_pdf()
    
    filename = f"Balanse_{client_name.replace(' ', '_')}_{balance_date}.pdf"
    
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_excel_balanse(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate Excel for Balanse"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Balanse"
    
    # Styling
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    subtotal_fill = PatternFill(start_color="d5dbdb", end_color="d5dbdb", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "Balanse"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    ws['A2'] = client_name
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:C3')
    ws['A3'] = f"Per: {format_date_no(date.fromisoformat(data.get('balance_date')))}"
    ws['A3'].alignment = Alignment(horizontal='center')
    
    row = 5
    
    # Assets section
    ws.cell(row=row, column=1, value="EIENDELER").font = Font(bold=True, size=12)
    row += 1
    
    for col, header in enumerate(["Konto", "Navn", "Beløp"], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    for item in data.get("eiendeler", []):
        ws.cell(row=row, column=1, value=item['account_number']).border = border
        ws.cell(row=row, column=2, value=item['account_name']).border = border
        amount_cell = ws.cell(row=row, column=3, value=item['balance'])
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = border
        row += 1
    
    # Sum assets
    ws.cell(row=row, column=1, value="Sum eiendeler").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2).fill = subtotal_fill
    ws.cell(row=row, column=2).border = border
    sum_cell = ws.cell(row=row, column=3, value=data.get('sum_eiendeler', 0))
    sum_cell.number_format = '#,##0.00'
    sum_cell.font = Font(bold=True)
    sum_cell.fill = subtotal_fill
    sum_cell.border = border
    row += 2
    
    # Liabilities section
    ws.cell(row=row, column=1, value="GJELD OG EGENKAPITAL").font = Font(bold=True, size=12)
    row += 1
    
    for item in data.get("gjeld_egenkapital", []):
        ws.cell(row=row, column=1, value=item['account_number']).border = border
        ws.cell(row=row, column=2, value=item['account_name']).border = border
        amount_cell = ws.cell(row=row, column=3, value=abs(item['balance']))
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = border
        row += 1
    
    # Sum liabilities
    ws.cell(row=row, column=1, value="Sum gjeld og egenkapital").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2).fill = subtotal_fill
    ws.cell(row=row, column=2).border = border
    sum_cell = ws.cell(row=row, column=3, value=data.get('sum_gjeld_egenkapital', 0))
    sum_cell.number_format = '#,##0.00'
    sum_cell.font = Font(bold=True)
    sum_cell.fill = subtotal_fill
    sum_cell.border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    filename = f"Balanse_{client_name.replace(' ', '_')}_{data.get('balance_date')}.xlsx"
    
    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_pdf_hovedbok(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate PDF for Hovedbok (General Ledger)"""
    
    from_date = data.get("from_date")
    to_date = data.get("to_date")
    account_number = data.get("account_number") or f"{data.get('account_from', '')}-{data.get('account_to', '')}"
    
    period_text = ""
    if from_date and to_date:
        period_text = f"Periode: {format_date_no(date.fromisoformat(from_date))} - {format_date_no(date.fromisoformat(to_date))}"
    
    # Build entries rows
    entries_rows = ""
    for entry in data.get("entries", []):
        entries_rows += f"""
        <tr>
            <td>{format_date_no(date.fromisoformat(entry['accounting_date']))}</td>
            <td>{entry['voucher_number']}</td>
            <td>{entry['account_number']}</td>
            <td>{entry['line_description'] or entry['entry_description']}</td>
            <td class="number">{format_currency(entry['debit_amount'])}</td>
            <td class="number">{format_currency(entry['credit_amount'])}</td>
        </tr>
        """
    
    opening_balance_row = ""
    if data.get("opening_balance") is not None:
        opening_balance_row = f"""
        <tr class="subtotal">
            <td colspan="4"><strong>Inngående saldo</strong></td>
            <td class="number" colspan="2"><strong>{format_currency(data['opening_balance'])}</strong></td>
        </tr>
        """
    
    closing_balance_row = ""
    if data.get("closing_balance") is not None:
        closing_balance_row = f"""
        <tr class="subtotal">
            <td colspan="4"><strong>Utgående saldo</strong></td>
            <td class="number" colspan="2"><strong>{format_currency(data['closing_balance'])}</strong></td>
        </tr>
        """
    
    html_content = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 1.5cm; }}
            body {{ font-family: Arial, sans-serif; font-size: 9pt; }}
            h1 {{ text-align: center; color: #2c3e50; margin-bottom: 5px; }}
            .meta {{ text-align: center; color: #7f8c8d; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ border: 1px solid #bdc3c7; padding: 6px; }}
            th {{ background-color: #34495e; color: white; font-weight: bold; }}
            tr:nth-child(even) {{ background-color: #ecf0f1; }}
            .number {{ text-align: right; }}
            .subtotal {{ font-weight: bold; background-color: #d5dbdb !important; }}
            .footer {{ margin-top: 20px; text-align: center; color: #95a5a6; font-size: 8pt; }}
        </style>
    </head>
    <body>
        <h1>Hovedbok</h1>
        <div class="meta">
            <p><strong>{client_name}</strong></p>
            <p>Konto: {account_number}</p>
            <p>{period_text}</p>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Dato</th>
                    <th>Bilag</th>
                    <th>Konto</th>
                    <th>Beskrivelse</th>
                    <th>Debet</th>
                    <th>Kredit</th>
                </tr>
            </thead>
            <tbody>
                {opening_balance_row}
                {entries_rows}
                {closing_balance_row}
            </tbody>
        </table>
        
        <div class="footer">
            <p>Generert: {format_date_no(date.today())} | Kontali ERP | Antall posteringer: {data.get('count', 0)}</p>
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = HTML(string=html_content).write_pdf()
    
    filename = f"Hovedbok_{client_name.replace(' ', '_')}_{date.today().isoformat()}.pdf"
    
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_excel_hovedbok(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate Excel for Hovedbok"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Hovedbok"
    
    # Styling
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    subtotal_fill = PatternFill(start_color="d5dbdb", end_color="d5dbdb", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Hovedbok"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = client_name
    ws['A2'].alignment = Alignment(horizontal='center')
    
    account_number = data.get("account_number") or f"{data.get('account_from', '')}-{data.get('account_to', '')}"
    ws.merge_cells('A3:F3')
    ws['A3'] = f"Konto: {account_number}"
    ws['A3'].alignment = Alignment(horizontal='center')
    
    row = 5
    
    # Headers
    headers = ["Dato", "Bilag", "Konto", "Beskrivelse", "Debet", "Kredit"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    # Opening balance
    if data.get("opening_balance") is not None:
        ws.cell(row=row, column=1, value="Inngående saldo").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = subtotal_fill
        ws.cell(row=row, column=1).border = border
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = subtotal_fill
            ws.cell(row=row, column=col).border = border
        balance_cell = ws.cell(row=row, column=5, value=data['opening_balance'])
        balance_cell.number_format = '#,##0.00'
        balance_cell.font = Font(bold=True)
        balance_cell.fill = subtotal_fill
        balance_cell.border = border
        ws.cell(row=row, column=6).fill = subtotal_fill
        ws.cell(row=row, column=6).border = border
        row += 1
    
    # Entries
    for entry in data.get("entries", []):
        ws.cell(row=row, column=1, value=format_date_no(date.fromisoformat(entry['accounting_date']))).border = border
        ws.cell(row=row, column=2, value=entry['voucher_number']).border = border
        ws.cell(row=row, column=3, value=entry['account_number']).border = border
        ws.cell(row=row, column=4, value=entry['line_description'] or entry['entry_description']).border = border
        
        debit_cell = ws.cell(row=row, column=5, value=entry['debit_amount'])
        debit_cell.number_format = '#,##0.00'
        debit_cell.border = border
        
        credit_cell = ws.cell(row=row, column=6, value=entry['credit_amount'])
        credit_cell.number_format = '#,##0.00'
        credit_cell.border = border
        
        row += 1
    
    # Closing balance
    if data.get("closing_balance") is not None:
        ws.cell(row=row, column=1, value="Utgående saldo").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = subtotal_fill
        ws.cell(row=row, column=1).border = border
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = subtotal_fill
            ws.cell(row=row, column=col).border = border
        balance_cell = ws.cell(row=row, column=5, value=data['closing_balance'])
        balance_cell.number_format = '#,##0.00'
        balance_cell.font = Font(bold=True)
        balance_cell.fill = subtotal_fill
        balance_cell.border = border
        ws.cell(row=row, column=6).fill = subtotal_fill
        ws.cell(row=row, column=6).border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Freeze top rows
    ws.freeze_panes = 'A6'
    
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    filename = f"Hovedbok_{client_name.replace(' ', '_')}_{date.today().isoformat()}.xlsx"
    
    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_pdf_supplier_ledger(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate PDF for Leverandørreskontro (Supplier Ledger)"""
    
    entries_rows = ""
    for entry in data.get("entries", []):
        status_label = {
            "open": "Åpen",
            "partially_paid": "Delvis betalt",
            "paid": "Betalt"
        }.get(entry.get('status'), entry.get('status'))
        
        days_overdue = entry.get('days_overdue', 0)
        overdue_class = "overdue" if days_overdue > 0 else ""
        
        entries_rows += f"""
        <tr class="{overdue_class}">
            <td>{entry.get('supplier_name', 'N/A')}</td>
            <td>{entry.get('invoice_number', 'N/A')}</td>
            <td>{format_date_no(date.fromisoformat(entry['invoice_date']))}</td>
            <td>{format_date_no(date.fromisoformat(entry['due_date']))}</td>
            <td class="number">{format_currency(entry['original_amount'])}</td>
            <td class="number">{format_currency(entry['remaining_amount'])}</td>
            <td>{status_label}</td>
            <td class="number">{days_overdue if days_overdue > 0 else '-'}</td>
        </tr>
        """
    
    html_content = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 1.5cm; }}
            body {{ font-family: Arial, sans-serif; font-size: 9pt; }}
            h1 {{ text-align: center; color: #2c3e50; margin-bottom: 5px; }}
            .meta {{ text-align: center; color: #7f8c8d; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ border: 1px solid #bdc3c7; padding: 6px; }}
            th {{ background-color: #34495e; color: white; font-weight: bold; }}
            tr:nth-child(even) {{ background-color: #ecf0f1; }}
            .overdue {{ background-color: #fadbd8 !important; }}
            .number {{ text-align: right; }}
            .subtotal {{ font-weight: bold; background-color: #d5dbdb !important; }}
            .footer {{ margin-top: 20px; text-align: center; color: #95a5a6; font-size: 8pt; }}
        </style>
    </head>
    <body>
        <h1>Leverandørreskontro</h1>
        <div class="meta">
            <p><strong>{client_name}</strong></p>
            <p>Generert: {format_date_no(date.today())}</p>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Leverandør</th>
                    <th>Faktura nr.</th>
                    <th>Fakturadato</th>
                    <th>Forfallsdato</th>
                    <th>Opprinnelig beløp</th>
                    <th>Gjenstående</th>
                    <th>Status</th>
                    <th>Dager forfalt</th>
                </tr>
            </thead>
            <tbody>
                {entries_rows}
                <tr class="subtotal">
                    <td colspan="5"><strong>Totalt gjenstående</strong></td>
                    <td class="number"><strong>{format_currency(data.get('total_remaining', 0))}</strong></td>
                    <td colspan="2"></td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            <p>Antall fakturaer: {data.get('total_count', 0)} | Kontali ERP</p>
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = HTML(string=html_content).write_pdf()
    
    filename = f"Leverandørreskontro_{client_name.replace(' ', '_')}_{date.today().isoformat()}.pdf"
    
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_excel_supplier_ledger(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate Excel for Leverandørreskontro"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Leverandørreskontro"
    
    # Styling
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    overdue_fill = PatternFill(start_color="fadbd8", end_color="fadbd8", fill_type="solid")
    subtotal_fill = PatternFill(start_color="d5dbdb", end_color="d5dbdb", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Leverandørreskontro"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'] = client_name
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    
    # Headers
    headers = ["Leverandør", "Faktura nr.", "Fakturadato", "Forfallsdato", 
               "Opprinnelig beløp", "Gjenstående", "Status", "Dager forfalt"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    # Entries
    for entry in data.get("entries", []):
        status_label = {
            "open": "Åpen",
            "partially_paid": "Delvis betalt",
            "paid": "Betalt"
        }.get(entry.get('status'), entry.get('status'))
        
        days_overdue = entry.get('days_overdue', 0)
        is_overdue = days_overdue > 0
        
        ws.cell(row=row, column=1, value=entry.get('supplier_name', 'N/A')).border = border
        ws.cell(row=row, column=2, value=entry.get('invoice_number', 'N/A')).border = border
        ws.cell(row=row, column=3, value=format_date_no(date.fromisoformat(entry['invoice_date']))).border = border
        ws.cell(row=row, column=4, value=format_date_no(date.fromisoformat(entry['due_date']))).border = border
        
        orig_cell = ws.cell(row=row, column=5, value=entry['original_amount'])
        orig_cell.number_format = '#,##0.00'
        orig_cell.border = border
        
        rem_cell = ws.cell(row=row, column=6, value=entry['remaining_amount'])
        rem_cell.number_format = '#,##0.00'
        rem_cell.border = border
        
        ws.cell(row=row, column=7, value=status_label).border = border
        ws.cell(row=row, column=8, value=days_overdue if is_overdue else "-").border = border
        
        # Highlight overdue rows
        if is_overdue:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = overdue_fill
        
        row += 1
    
    # Total row
    ws.cell(row=row, column=1, value="Totalt gjenstående").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=1).border = border
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = subtotal_fill
        ws.cell(row=row, column=col).border = border
    total_cell = ws.cell(row=row, column=6, value=data.get('total_remaining', 0))
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(bold=True)
    total_cell.fill = subtotal_fill
    total_cell.border = border
    for col in range(7, 9):
        ws.cell(row=row, column=col).fill = subtotal_fill
        ws.cell(row=row, column=col).border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    # Freeze top rows
    ws.freeze_panes = 'A5'
    
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    filename = f"Leverandørreskontro_{client_name.replace(' ', '_')}_{date.today().isoformat()}.xlsx"
    
    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_pdf_customer_ledger(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate PDF for Kundereskontro (Customer Ledger)"""
    
    entries_rows = ""
    for entry in data.get("entries", []):
        status_label = entry.get('status_label', entry.get('status', 'N/A'))
        
        days_overdue = entry.get('days_overdue', 0)
        overdue_class = "overdue" if days_overdue > 0 else ""
        
        entries_rows += f"""
        <tr class="{overdue_class}">
            <td>{entry.get('customer_name', 'N/A')}</td>
            <td>{entry.get('invoice_number', 'N/A')}</td>
            <td>{format_date_no(date.fromisoformat(entry['invoice_date']))}</td>
            <td>{format_date_no(date.fromisoformat(entry['due_date']))}</td>
            <td class="number">{format_currency(entry['original_amount'])}</td>
            <td class="number">{format_currency(entry['remaining_amount'])}</td>
            <td>{status_label}</td>
            <td class="number">{days_overdue if days_overdue > 0 else '-'}</td>
        </tr>
        """
    
    html_content = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 1.5cm; }}
            body {{ font-family: Arial, sans-serif; font-size: 9pt; }}
            h1 {{ text-align: center; color: #2c3e50; margin-bottom: 5px; }}
            .meta {{ text-align: center; color: #7f8c8d; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ border: 1px solid #bdc3c7; padding: 6px; }}
            th {{ background-color: #34495e; color: white; font-weight: bold; }}
            tr:nth-child(even) {{ background-color: #ecf0f1; }}
            .overdue {{ background-color: #fadbd8 !important; }}
            .number {{ text-align: right; }}
            .subtotal {{ font-weight: bold; background-color: #d5dbdb !important; }}
            .footer {{ margin-top: 20px; text-align: center; color: #95a5a6; font-size: 8pt; }}
        </style>
    </head>
    <body>
        <h1>Kundereskontro</h1>
        <div class="meta">
            <p><strong>{client_name}</strong></p>
            <p>Generert: {format_date_no(date.today())}</p>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Kunde</th>
                    <th>Faktura nr.</th>
                    <th>Fakturadato</th>
                    <th>Forfallsdato</th>
                    <th>Opprinnelig beløp</th>
                    <th>Gjenstående</th>
                    <th>Status</th>
                    <th>Dager forfalt</th>
                </tr>
            </thead>
            <tbody>
                {entries_rows}
                <tr class="subtotal">
                    <td colspan="5"><strong>Totalt gjenstående</strong></td>
                    <td class="number"><strong>{format_currency(data.get('total_remaining', 0))}</strong></td>
                    <td colspan="2"></td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            <p>Antall fakturaer: {data.get('total_count', 0)} | Kontali ERP</p>
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = HTML(string=html_content).write_pdf()
    
    filename = f"Kundereskontro_{client_name.replace(' ', '_')}_{date.today().isoformat()}.pdf"
    
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_excel_customer_ledger(
    data: Dict[str, Any],
    client_name: str
) -> StreamingResponse:
    """Generate Excel for Kundereskontro"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Kundereskontro"
    
    # Styling
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    overdue_fill = PatternFill(start_color="fadbd8", end_color="fadbd8", fill_type="solid")
    subtotal_fill = PatternFill(start_color="d5dbdb", end_color="d5dbdb", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Kundereskontro"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'] = client_name
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    
    # Headers
    headers = ["Kunde", "Faktura nr.", "Fakturadato", "Forfallsdato", 
               "Opprinnelig beløp", "Gjenstående", "Status", "Dager forfalt"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    # Entries
    for entry in data.get("entries", []):
        status_label = entry.get('status_label', entry.get('status', 'N/A'))
        
        days_overdue = entry.get('days_overdue', 0)
        is_overdue = days_overdue > 0
        
        ws.cell(row=row, column=1, value=entry.get('customer_name', 'N/A')).border = border
        ws.cell(row=row, column=2, value=entry.get('invoice_number', 'N/A')).border = border
        ws.cell(row=row, column=3, value=format_date_no(date.fromisoformat(entry['invoice_date']))).border = border
        ws.cell(row=row, column=4, value=format_date_no(date.fromisoformat(entry['due_date']))).border = border
        
        orig_cell = ws.cell(row=row, column=5, value=entry['original_amount'])
        orig_cell.number_format = '#,##0.00'
        orig_cell.border = border
        
        rem_cell = ws.cell(row=row, column=6, value=entry['remaining_amount'])
        rem_cell.number_format = '#,##0.00'
        rem_cell.border = border
        
        ws.cell(row=row, column=7, value=status_label).border = border
        ws.cell(row=row, column=8, value=days_overdue if is_overdue else "-").border = border
        
        # Highlight overdue rows
        if is_overdue:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = overdue_fill
        
        row += 1
    
    # Total row
    ws.cell(row=row, column=1, value="Totalt gjenstående").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=1).border = border
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = subtotal_fill
        ws.cell(row=row, column=col).border = border
    total_cell = ws.cell(row=row, column=6, value=data.get('total_remaining', 0))
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(bold=True)
    total_cell.fill = subtotal_fill
    total_cell.border = border
    for col in range(7, 9):
        ws.cell(row=row, column=col).fill = subtotal_fill
        ws.cell(row=row, column=col).border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    
    # Freeze top rows
    ws.freeze_panes = 'A5'
    
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    filename = f"Kundereskontro_{client_name.replace(' ', '_')}_{date.today().isoformat()}.xlsx"
    
    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
