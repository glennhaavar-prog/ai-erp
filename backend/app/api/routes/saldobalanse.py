"""
Saldobalanse Report API - Trial Balance
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import date, datetime
from typing import Optional, Dict, Any
from uuid import UUID
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas

from app.database import get_db
from app.services.report_service import calculate_saldobalanse, get_saldobalanse_summary

router = APIRouter(prefix="/api/reports/saldobalanse", tags=["Reports - Saldobalanse"])


@router.get("/")
async def get_saldobalanse(
    client_id: UUID = Query(..., description="Client ID to filter accounts"),
    from_date: Optional[date] = Query(None, description="Start date for transactions (optional)"),
    to_date: Optional[date] = Query(None, description="End date for transactions (optional)"),
    account_class: Optional[str] = Query(None, description="Filter by account class (e.g., '1', '2', '3')"),
    include_summary: bool = Query(True, description="Include summary statistics"),
    db: AsyncSession = Depends(get_db)
) -> Dict[str, Any]:
    """
    Get Saldobalanse (Trial Balance) report.
    
    Returns balance information for all accounts:
    - Opening balance (inngående saldo)
    - Total debit/credit transactions
    - Current balance (nåværende saldo)
    
    Filters:
    - from_date/to_date: Date range for transactions
    - account_class: Filter by first digit(s) of account number (e.g., "1" for assets)
    
    Returns:
    - accounts: List of accounts with balance data
    - summary: Aggregated totals (if include_summary=true)
    - filters: Applied filters
    - timestamp: Report generation time
    """
    
    # Validate account_class if provided
    if account_class and not account_class.isdigit():
        raise HTTPException(
            status_code=400,
            detail="account_class must be numeric (e.g., '1', '2', '15')"
        )
    
    # Calculate saldobalanse
    accounts = await calculate_saldobalanse(
        db=db,
        client_id=client_id,
        from_date=from_date,
        to_date=to_date,
        account_class=account_class
    )
    
    # Build response
    response = {
        "accounts": accounts,
        "filters": {
            "client_id": str(client_id),
            "from_date": from_date.isoformat() if from_date else None,
            "to_date": to_date.isoformat() if to_date else None,
            "account_class": account_class,
        },
        "timestamp": datetime.utcnow().isoformat(),
    }
    
    # Add summary if requested
    if include_summary:
        summary = await get_saldobalanse_summary(accounts)
        response["summary"] = summary
    
    return response


@router.get("/export/excel/")
async def export_saldobalanse_excel(
    client_id: UUID = Query(..., description="Client ID to filter accounts"),
    from_date: Optional[date] = Query(None, description="Start date for transactions"),
    to_date: Optional[date] = Query(None, description="End date for transactions"),
    account_class: Optional[str] = Query(None, description="Filter by account class"),
    db: AsyncSession = Depends(get_db)
):
    """
    Export Saldobalanse to Excel (.xlsx format).
    
    Returns a downloadable Excel file with trial balance data.
    """
    
    # Validate account_class if provided
    if account_class and not account_class.isdigit():
        raise HTTPException(
            status_code=400,
            detail="account_class must be numeric"
        )
    
    # Calculate saldobalanse
    accounts = await calculate_saldobalanse(
        db=db,
        client_id=client_id,
        from_date=from_date,
        to_date=to_date,
        account_class=account_class
    )
    
    # Get summary
    summary = await get_saldobalanse_summary(accounts)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Saldobalanse"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "SALDOBALANSERAPPORT"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Metadata
    row = 2
    ws[f'A{row}'] = "Klient ID:"
    ws[f'B{row}'] = str(client_id)
    row += 1
    
    if from_date:
        ws[f'A{row}'] = "Fra dato:"
        ws[f'B{row}'] = from_date.strftime("%Y-%m-%d")
        row += 1
    
    if to_date:
        ws[f'A{row}'] = "Til dato:"
        ws[f'B{row}'] = to_date.strftime("%Y-%m-%d")
        row += 1
    
    ws[f'A{row}'] = "Generert:"
    ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row += 2
    
    # Column headers
    headers = [
        "Kontonr",
        "Kontonavn",
        "Type",
        "Inngående saldo",
        "Debet",
        "Kredit",
        "Endring",
        "Nåværende saldo"
    ]
    
    header_row = row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row += 1
    
    # Data rows
    for account in accounts:
        ws.cell(row=row, column=1, value=account["account_number"])
        ws.cell(row=row, column=2, value=account["account_name"])
        ws.cell(row=row, column=3, value=account["account_type"])
        ws.cell(row=row, column=4, value=account["opening_balance"])
        ws.cell(row=row, column=5, value=account["total_debit"])
        ws.cell(row=row, column=6, value=account["total_credit"])
        ws.cell(row=row, column=7, value=account["net_change"])
        ws.cell(row=row, column=8, value=account["current_balance"])
        
        # Format numbers
        for col in [4, 5, 6, 7, 8]:
            cell = ws.cell(row=row, column=col)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            cell.border = border
        
        # Border for text cells
        for col in [1, 2, 3]:
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Summary section
    row += 1
    summary_start = row
    
    # Summary header
    ws.merge_cells(f'A{row}:C{row}')
    summary_header = ws.cell(row=row, column=1)
    summary_header.value = "SAMMENDRAG"
    summary_header.font = Font(bold=True, size=12)
    row += 1
    
    # Total accounts
    ws.cell(row=row, column=1, value="Totalt antall kontoer:")
    ws.cell(row=row, column=2, value=summary["total_accounts"])
    row += 1
    
    # Total debit
    ws.cell(row=row, column=1, value="Sum debet:")
    ws.cell(row=row, column=2, value=summary["total_debit"])
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    row += 1
    
    # Total credit
    ws.cell(row=row, column=1, value="Sum kredit:")
    ws.cell(row=row, column=2, value=summary["total_credit"])
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    row += 1
    
    # Balance check
    ws.cell(row=row, column=1, value="Balansert:")
    ws.cell(row=row, column=2, value="JA" if summary["balance_check"]["balanced"] else "NEI")
    if not summary["balance_check"]["balanced"]:
        ws.cell(row=row, column=2).font = Font(bold=True, color="FF0000")
        ws.cell(row=row, column=3, value=f"Differanse: {summary['balance_check']['difference']}")
    row += 1
    
    # By type breakdown
    row += 1
    ws.cell(row=row, column=1, value="Per kontotype:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    for account_type, type_data in summary["by_type"].items():
        ws.cell(row=row, column=1, value=f"  {account_type.upper()}:")
        ws.cell(row=row, column=2, value=f"{type_data['count']} kontoer")
        ws.cell(row=row, column=3, value=f"Saldo: {type_data['current_balance']:.2f}")
        row += 1
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generate filename
    filename = f"saldobalanse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Return as streaming response
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/pdf/")
async def export_saldobalanse_pdf(
    client_id: UUID = Query(..., description="Client ID to filter accounts"),
    from_date: Optional[date] = Query(None, description="Start date for transactions"),
    to_date: Optional[date] = Query(None, description="End date for transactions"),
    account_class: Optional[str] = Query(None, description="Filter by account class"),
    db: AsyncSession = Depends(get_db)
):
    """
    Export Saldobalanse to PDF format.
    
    Returns a downloadable PDF file with trial balance data.
    """
    
    # Validate account_class if provided
    if account_class and not account_class.isdigit():
        raise HTTPException(
            status_code=400,
            detail="account_class must be numeric"
        )
    
    # Calculate saldobalanse
    accounts = await calculate_saldobalanse(
        db=db,
        client_id=client_id,
        from_date=from_date,
        to_date=to_date,
        account_class=account_class
    )
    
    # Get summary
    summary = await get_saldobalanse_summary(accounts)
    
    # Create PDF
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm,
    )
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph("SALDOBALANSERAPPORT", title_style)
    elements.append(title)
    
    # Metadata
    meta_data = [
        ["Klient ID:", str(client_id)],
        ["Fra dato:", from_date.strftime("%Y-%m-%d") if from_date else "N/A"],
        ["Til dato:", to_date.strftime("%Y-%m-%d") if to_date else "N/A"],
        ["Generert:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    
    meta_table = Table(meta_data, colWidths=[40*mm, 80*mm])
    meta_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 20))
    
    # Data table
    table_data = [[
        "Kontonr",
        "Kontonavn",
        "Type",
        "Inng. saldo",
        "Debet",
        "Kredit",
        "Nåv. saldo"
    ]]
    
    for account in accounts:
        table_data.append([
            account["account_number"],
            account["account_name"][:30],  # Truncate long names
            account["account_type"][:10],
            f"{account['opening_balance']:.2f}",
            f"{account['total_debit']:.2f}",
            f"{account['total_credit']:.2f}",
            f"{account['current_balance']:.2f}",
        ])
    
    # Create table
    data_table = Table(table_data, colWidths=[
        20*mm, 50*mm, 25*mm, 25*mm, 25*mm, 25*mm, 25*mm
    ])
    
    data_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Right-align numbers
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
    ]))
    
    elements.append(data_table)
    elements.append(Spacer(1, 20))
    
    # Summary section
    summary_title = Paragraph("<b>SAMMENDRAG</b>", styles['Heading2'])
    elements.append(summary_title)
    
    summary_data = [
        ["Totalt antall kontoer:", str(summary["total_accounts"])],
        ["Sum debet:", f"{summary['total_debit']:.2f}"],
        ["Sum kredit:", f"{summary['total_credit']:.2f}"],
        ["Balansert:", "JA" if summary["balance_check"]["balanced"] else f"NEI (diff: {summary['balance_check']['difference']:.2f})"],
    ]
    
    summary_table = Table(summary_data, colWidths=[60*mm, 60*mm])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    
    # Build PDF
    doc.build(elements)
    
    pdf_buffer.seek(0)
    
    # Generate filename
    filename = f"saldobalanse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    # Return as streaming response
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
