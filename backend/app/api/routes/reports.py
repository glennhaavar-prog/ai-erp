"""
Reports API - Hovedbok (General Ledger) and other reports
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func
from sqlalchemy.orm import selectinload, joinedload
from datetime import date, datetime
from typing import Optional, List, Dict, Any
from uuid import UUID
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from app.database import get_db
from app.models.general_ledger import GeneralLedger, GeneralLedgerLine
from app.models.vendor_invoice import VendorInvoice
from app.models.vendor import Vendor

router = APIRouter(prefix="/api/reports", tags=["Reports"])


@router.get("/hovedbok/")
async def get_hovedbok(
    client_id: UUID = Query(..., description="Client ID to filter entries"),
    start_date: Optional[date] = Query(None, description="Start date (accounting_date >= start_date)"),
    end_date: Optional[date] = Query(None, description="End date (accounting_date <= end_date)"),
    account_number: Optional[str] = Query(None, description="Filter by account number"),
    vendor_id: Optional[UUID] = Query(None, description="Filter by vendor ID"),
    status: Optional[str] = Query(None, description="Filter by status (posted/reversed)"),
    voucher_series: Optional[str] = Query(None, description="Filter by voucher series (A/B/C)"),
    sort_by: str = Query("accounting_date", description="Sort field (accounting_date, voucher_number, created_at)"),
    sort_order: str = Query("asc", description="Sort order (asc/desc)"),
    page: int = Query(1, ge=1, description="Page number (starts at 1)"),
    page_size: int = Query(50, ge=1, le=500, description="Number of entries per page"),
    include_lines: bool = Query(True, description="Include general ledger lines in response"),
    include_invoice: bool = Query(True, description="Include vendor invoice details in response"),
    db: AsyncSession = Depends(get_db)
) -> Dict[str, Any]:
    """
    Get Hovedbok (General Ledger) entries with filtering, sorting, and pagination.
    
    This endpoint joins GeneralLedger + GeneralLedgerLine + VendorInvoice to provide
    complete bookkeeping data.
    
    Returns:
    - entries: List of general ledger entries with lines and invoice details
    - pagination: Page metadata
    - summary: Aggregated totals for the filtered data
    """
    
    # Build base query
    query = select(GeneralLedger).where(GeneralLedger.client_id == client_id)
    
    # Apply filters
    filters = []
    
    if start_date:
        filters.append(GeneralLedger.accounting_date >= start_date)
    
    if end_date:
        filters.append(GeneralLedger.accounting_date <= end_date)
    
    if status:
        filters.append(GeneralLedger.status == status)
    
    if voucher_series:
        filters.append(GeneralLedger.voucher_series == voucher_series)
    
    # Filter by account number (need to join lines)
    if account_number:
        query = query.join(GeneralLedgerLine)
        filters.append(GeneralLedgerLine.account_number == account_number)
    
    # Filter by vendor (need to join invoice)
    if vendor_id:
        query = query.join(
            VendorInvoice,
            and_(
                VendorInvoice.general_ledger_id == GeneralLedger.id,
                VendorInvoice.vendor_id == vendor_id
            )
        )
    
    if filters:
        query = query.where(and_(*filters))
    
    # Eager load relationships if requested
    if include_lines:
        query = query.options(selectinload(GeneralLedger.lines))
    
    # Apply sorting
    sort_field_map = {
        "accounting_date": GeneralLedger.accounting_date,
        "voucher_number": GeneralLedger.voucher_number,
        "created_at": GeneralLedger.created_at,
        "entry_date": GeneralLedger.entry_date,
    }
    
    sort_field = sort_field_map.get(sort_by, GeneralLedger.accounting_date)
    
    if sort_order.lower() == "desc":
        query = query.order_by(sort_field.desc())
    else:
        query = query.order_by(sort_field.asc())
    
    # Add secondary sort by voucher_number for consistency
    query = query.order_by(GeneralLedger.voucher_number.asc())
    
    # Get total count (before pagination)
    count_query = select(func.count()).select_from(query.distinct().subquery())
    total_result = await db.execute(count_query)
    total_count = total_result.scalar() or 0
    
    # Apply pagination
    offset = (page - 1) * page_size
    query = query.limit(page_size).offset(offset)
    
    # Execute query
    result = await db.execute(query)
    entries = result.scalars().unique().all()
    
    # Build response data
    response_entries = []
    
    for entry in entries:
        entry_dict = {
            "id": str(entry.id),
            "client_id": str(entry.client_id),
            "voucher_number": entry.voucher_number,
            "voucher_series": entry.voucher_series,
            "full_voucher": f"{entry.voucher_series}-{entry.voucher_number}",
            "accounting_date": entry.accounting_date.isoformat(),
            "entry_date": entry.entry_date.isoformat(),
            "period": entry.period,
            "fiscal_year": entry.fiscal_year,
            "description": entry.description,
            "source_type": entry.source_type,
            "source_id": str(entry.source_id) if entry.source_id else None,
            "created_by_type": entry.created_by_type,
            "created_by_id": str(entry.created_by_id) if entry.created_by_id else None,
            "status": entry.status,
            "is_reversed": entry.is_reversed,
            "reversed_by_entry_id": str(entry.reversed_by_entry_id) if entry.reversed_by_entry_id else None,
            "reversal_reason": entry.reversal_reason,
            "locked": entry.locked,
            "created_at": entry.created_at.isoformat(),
        }
        
        # Add lines if requested
        if include_lines and entry.lines:
            entry_dict["lines"] = []
            total_debit = Decimal("0.00")
            total_credit = Decimal("0.00")
            
            for line in sorted(entry.lines, key=lambda l: l.line_number):
                line_dict = {
                    "id": str(line.id),
                    "line_number": line.line_number,
                    "account_number": line.account_number,
                    "debit_amount": float(line.debit_amount),
                    "credit_amount": float(line.credit_amount),
                    "vat_code": line.vat_code,
                    "vat_amount": float(line.vat_amount) if line.vat_amount else 0.0,
                    "vat_base_amount": float(line.vat_base_amount) if line.vat_base_amount else None,
                    "line_description": line.line_description,
                    "department_id": str(line.department_id) if line.department_id else None,
                    "project_id": str(line.project_id) if line.project_id else None,
                    "cost_center_id": str(line.cost_center_id) if line.cost_center_id else None,
                    "ai_confidence_score": line.ai_confidence_score,
                    "ai_reasoning": line.ai_reasoning,
                }
                entry_dict["lines"].append(line_dict)
                total_debit += line.debit_amount
                total_credit += line.credit_amount
            
            # Add totals for verification
            entry_dict["totals"] = {
                "debit": float(total_debit),
                "credit": float(total_credit),
                "balanced": abs(total_debit - total_credit) < Decimal("0.01")
            }
        
        # Add vendor invoice details if requested and available
        if include_invoice and entry.source_type == "ehf_invoice" and entry.source_id:
            # Query the invoice
            invoice_query = select(VendorInvoice).where(
                VendorInvoice.id == entry.source_id
            ).options(joinedload(VendorInvoice.vendor))
            
            invoice_result = await db.execute(invoice_query)
            invoice = invoice_result.scalar_one_or_none()
            
            if invoice:
                invoice_dict = {
                    "id": str(invoice.id),
                    "invoice_number": invoice.invoice_number,
                    "invoice_date": invoice.invoice_date.isoformat(),
                    "due_date": invoice.due_date.isoformat(),
                    "amount_excl_vat": float(invoice.amount_excl_vat),
                    "vat_amount": float(invoice.vat_amount),
                    "total_amount": float(invoice.total_amount),
                    "currency": invoice.currency,
                    "payment_status": invoice.payment_status,
                    "vendor": None
                }
                
                if invoice.vendor:
                    invoice_dict["vendor"] = {
                        "id": str(invoice.vendor.id),
                        "name": invoice.vendor.name,
                        "org_number": invoice.vendor.org_number,
                        "vendor_number": invoice.vendor.vendor_number,
                    }
                
                entry_dict["invoice"] = invoice_dict
        
        response_entries.append(entry_dict)
    
    # Calculate pagination metadata
    total_pages = (total_count + page_size - 1) // page_size  # Ceiling division
    has_next = page < total_pages
    has_prev = page > 1
    
    pagination = {
        "page": page,
        "page_size": page_size,
        "total_entries": total_count,
        "total_pages": total_pages,
        "has_next": has_next,
        "has_prev": has_prev,
    }
    
    # Calculate summary statistics for filtered data
    # Sum up all debits and credits from the lines
    summary_query = select(
        func.sum(GeneralLedgerLine.debit_amount).label("total_debit"),
        func.sum(GeneralLedgerLine.credit_amount).label("total_credit"),
        func.count(func.distinct(GeneralLedger.id)).label("entry_count"),
    ).select_from(GeneralLedger).join(GeneralLedgerLine).where(
        GeneralLedger.client_id == client_id
    )
    
    # Apply same filters for summary
    if start_date:
        summary_query = summary_query.where(GeneralLedger.accounting_date >= start_date)
    if end_date:
        summary_query = summary_query.where(GeneralLedger.accounting_date <= end_date)
    if status:
        summary_query = summary_query.where(GeneralLedger.status == status)
    if voucher_series:
        summary_query = summary_query.where(GeneralLedger.voucher_series == voucher_series)
    if account_number:
        summary_query = summary_query.where(GeneralLedgerLine.account_number == account_number)
    
    summary_result = await db.execute(summary_query)
    summary_row = summary_result.one()
    
    summary = {
        "total_debit": float(summary_row.total_debit or 0),
        "total_credit": float(summary_row.total_credit or 0),
        "total_entries": summary_row.entry_count or 0,
        "date_range": {
            "start": start_date.isoformat() if start_date else None,
            "end": end_date.isoformat() if end_date else None,
        },
        "filters_applied": {
            "account_number": account_number,
            "vendor_id": str(vendor_id) if vendor_id else None,
            "status": status,
            "voucher_series": voucher_series,
        }
    }
    
    return {
        "entries": response_entries,
        "pagination": pagination,
        "summary": summary,
        "timestamp": datetime.utcnow().isoformat(),
    }


@router.get("/hovedbok/{entry_id}")
async def get_hovedbok_entry(
    entry_id: UUID,
    db: AsyncSession = Depends(get_db)
) -> Dict[str, Any]:
    """
    Get a single Hovedbok entry by ID with all details.
    """
    query = select(GeneralLedger).where(
        GeneralLedger.id == entry_id
    ).options(
        selectinload(GeneralLedger.lines)
    )
    
    result = await db.execute(query)
    entry = result.scalar_one_or_none()
    
    if not entry:
        raise HTTPException(status_code=404, detail="General ledger entry not found")
    
    # Build detailed response
    entry_dict = {
        "id": str(entry.id),
        "client_id": str(entry.client_id),
        "voucher_number": entry.voucher_number,
        "voucher_series": entry.voucher_series,
        "full_voucher": f"{entry.voucher_series}-{entry.voucher_number}",
        "accounting_date": entry.accounting_date.isoformat(),
        "entry_date": entry.entry_date.isoformat(),
        "period": entry.period,
        "fiscal_year": entry.fiscal_year,
        "description": entry.description,
        "source_type": entry.source_type,
        "source_id": str(entry.source_id) if entry.source_id else None,
        "created_by_type": entry.created_by_type,
        "created_by_id": str(entry.created_by_id) if entry.created_by_id else None,
        "status": entry.status,
        "is_reversed": entry.is_reversed,
        "reversed_by_entry_id": str(entry.reversed_by_entry_id) if entry.reversed_by_entry_id else None,
        "reversal_reason": entry.reversal_reason,
        "locked": entry.locked,
        "created_at": entry.created_at.isoformat(),
        "lines": []
    }
    
    # Add lines
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")
    
    for line in sorted(entry.lines, key=lambda l: l.line_number):
        line_dict = {
            "id": str(line.id),
            "line_number": line.line_number,
            "account_number": line.account_number,
            "debit_amount": float(line.debit_amount),
            "credit_amount": float(line.credit_amount),
            "vat_code": line.vat_code,
            "vat_amount": float(line.vat_amount) if line.vat_amount else 0.0,
            "vat_base_amount": float(line.vat_base_amount) if line.vat_base_amount else None,
            "line_description": line.line_description,
            "department_id": str(line.department_id) if line.department_id else None,
            "project_id": str(line.project_id) if line.project_id else None,
            "cost_center_id": str(line.cost_center_id) if line.cost_center_id else None,
            "ai_confidence_score": line.ai_confidence_score,
            "ai_reasoning": line.ai_reasoning,
        }
        entry_dict["lines"].append(line_dict)
        total_debit += line.debit_amount
        total_credit += line.credit_amount
    
    # Add totals
    entry_dict["totals"] = {
        "debit": float(total_debit),
        "credit": float(total_credit),
        "balanced": abs(total_debit - total_credit) < Decimal("0.01")
    }
    
    # Add invoice details if available
    if entry.source_type == "ehf_invoice" and entry.source_id:
        invoice_query = select(VendorInvoice).where(
            VendorInvoice.id == entry.source_id
        ).options(joinedload(VendorInvoice.vendor))
        
        invoice_result = await db.execute(invoice_query)
        invoice = invoice_result.scalar_one_or_none()
        
        if invoice:
            invoice_dict = {
                "id": str(invoice.id),
                "invoice_number": invoice.invoice_number,
                "invoice_date": invoice.invoice_date.isoformat(),
                "due_date": invoice.due_date.isoformat(),
                "amount_excl_vat": float(invoice.amount_excl_vat),
                "vat_amount": float(invoice.vat_amount),
                "total_amount": float(invoice.total_amount),
                "currency": invoice.currency,
                "payment_status": invoice.payment_status,
                "ai_confidence_score": invoice.ai_confidence_score,
                "ai_detected_category": invoice.ai_detected_category,
                "vendor": None
            }
            
            if invoice.vendor:
                invoice_dict["vendor"] = {
                    "id": str(invoice.vendor.id),
                    "name": invoice.vendor.name,
                    "org_number": invoice.vendor.org_number,
                    "vendor_number": invoice.vendor.vendor_number,
                    "account_number": invoice.vendor.account_number,
                }
            
            entry_dict["invoice"] = invoice_dict
    
    return entry_dict


@router.get("/hovedbok/export/")
async def export_hovedbok_to_excel(
    client_id: UUID = Query(..., description="Client ID to filter entries"),
    start_date: Optional[date] = Query(None, description="Start date"),
    end_date: Optional[date] = Query(None, description="End date"),
    account_number: Optional[str] = Query(None, description="Filter by account number"),
    vendor_id: Optional[UUID] = Query(None, description="Filter by vendor ID"),
    status: Optional[str] = Query(None, description="Filter by status (posted/reversed)"),
    db: AsyncSession = Depends(get_db)
):
    """
    Export Hovedbok (General Ledger) to Excel (.xlsx format)
    
    Returns a downloadable Excel file with all GL entries matching the filters.
    """
    
    # Build query (same logic as GET endpoint)
    query = select(GeneralLedger).where(GeneralLedger.client_id == client_id)
    
    # Apply filters
    if start_date:
        query = query.where(GeneralLedger.accounting_date >= start_date)
    if end_date:
        query = query.where(GeneralLedger.accounting_date <= end_date)
    if status:
        query = query.where(GeneralLedger.status == status)
    
    # Load with relationships
    query = query.options(
        selectinload(GeneralLedger.lines),
        joinedload(GeneralLedger.vendor)
    )
    
    # Sort by accounting date ascending
    query = query.order_by(GeneralLedger.accounting_date.asc(), GeneralLedger.created_at.asc())
    
    # Execute query
    result = await db.execute(query)
    entries = result.scalars().unique().all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Hovedbok"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Bilagsnr",
        "Dato",
        "Beskrivelse",
        "Konto",
        "Debet",
        "Kredit",
        "Saldo",
        "Leverandør",
        "Status"
    ]
    
    ws.append(headers)
    
    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    running_balance = Decimal("0.00")
    row_num = 2
    
    for entry in entries:
        # For each general ledger line
        for line in entry.lines:
            debit = float(line.debit_amount) if line.debit_amount else 0
            credit = float(line.credit_amount) if line.credit_amount else 0
            
            if entry.status != "reversed":
                running_balance += Decimal(str(debit)) - Decimal(str(credit))
            
            vendor_name = entry.vendor.name if entry.vendor else ""
            status_text = "Postert" if entry.status == "posted" else "Reversert"
            
            row_data = [
                entry.voucher_number,
                entry.accounting_date.strftime("%Y-%m-%d"),
                line.description or entry.description,
                line.account_number,
                debit if debit != 0 else "",
                credit if credit != 0 else "",
                float(running_balance),
                vendor_name,
                status_text
            ]
            
            ws.append(row_data)
            
            # Style data row
            for col_num in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                
                # Align numbers right
                if col_num in [5, 6, 7]:  # Debet, Kredit, Saldo
                    cell.alignment = Alignment(horizontal="right")
                    # Format as currency
                    if cell.value != "":
                        cell.number_format = '#,##0.00'
            
            row_num += 1
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Add summary row
    ws.append([])
    summary_row = row_num + 1
    ws.cell(row=summary_row, column=1, value="TOTALER:")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    # Calculate totals
    total_debit = sum(float(line.debit_amount) for entry in entries for line in entry.lines if entry.status != "reversed")
    total_credit = sum(float(line.credit_amount) for entry in entries for line in entry.lines if entry.status != "reversed")
    
    ws.cell(row=summary_row, column=5, value=total_debit).font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=total_credit).font = Font(bold=True)
    ws.cell(row=summary_row, column=5).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=6).number_format = '#,##0.00'
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generate filename
    filename = f"hovedbok_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Return as streaming response
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
